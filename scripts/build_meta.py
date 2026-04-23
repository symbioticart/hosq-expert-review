#!/usr/bin/env python3
"""
Build data/<projectId>/meta.json for 5 projects by reading one row per project
from hosq_project_profile_v7_enriched.xlsx.

Headers live on row 4 (column_id). Display labels live on row 2.
Project rows: 17 (byob), 26 (notations-lab), 32 (art-science), 37 (jam-discipline), 38 (cosmic).
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "hosq_project_profile_v7_enriched.xlsx"

ROW_TO_PROJECT = {
    17: "byob",
    26: "notations-lab",
    32: "art-science",
    37: "jam-discipline",
    38: "cosmic",
}

# UI is English-only — override Russian labels from the spreadsheet with EN equivalents.
EN_LABELS = {
    "short_description":              "Short description",
    "full_description":               "Full description",
    "curatorial_concept":             "Curatorial concept",
    "artistic_research_questions":    "Research questions",

    "project_type":                   "Project type",
    "project_edition_year":           "Edition year",
    "project_status":                 "Status",
    "hosq_stream":                    "HOSQ stream",
    "project_format_tags":            "Format tags",
    "start_date":                     "Start date",
    "end_date":                       "End date",
    "duration_days":                  "Duration (days)",
    "primary_city":                   "City",
    "primary_country":                "Country",
    "venue_name_primary":             "Primary venue",
    "venue_type":                     "Venue type",
    "team_size_total":                "Team size",
    "participants_count_onsite":      "Active participants",
    "partners_international_count":   "International partners",
    "partners_local_count":           "Local partners",
    "jobs_created_count":             "Jobs created",
    "jobs_paid_count":                "Paid positions",
    "budget_total_usd":               "Total budget (USD)",
    "target_audience_description":    "Target audience",
    "languages_offered":              "Languages",
    "audience_reach_onsite":          "On-site audience",
    "audience_reach_online":          "Online audience",
}

# Fields surfaced in the "Description" markdown block (in this order).
DESCRIPTION_FIELDS = [
    "short_description",
    "full_description",
    "curatorial_concept",
    "artistic_research_questions",
]

# Fields surfaced as rows in the key-facts Data table.
TABLE_FIELDS = [
    "project_type", "project_edition_year", "project_status", "hosq_stream",
    "project_format_tags", "start_date", "end_date", "duration_days",
    "primary_city", "primary_country", "venue_name_primary", "venue_type",
    "team_size_total", "participants_count_onsite",
    "partners_international_count", "partners_local_count",
    "jobs_created_count", "jobs_paid_count", "budget_total_usd",
    "target_audience_description", "languages_offered",
    "audience_reach_onsite", "audience_reach_online",
]


def main() -> int:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["projects"]

    cid_row = 4
    col_ids = {ws.cell(cid_row, c).value: c for c in range(2, ws.max_column + 1)
               if ws.cell(cid_row, c).value}

    def label_for(field: str) -> str:
        return EN_LABELS.get(field, field.replace("_", " ").capitalize())

    for row_num, pid in ROW_TO_PROJECT.items():
        def v(field: str):
            c = col_ids.get(field)
            return ws.cell(row_num, c).value if c else None

        name = v("project_title") or pid
        parts = []
        for f in DESCRIPTION_FIELDS:
            txt = v(f)
            if txt is None or str(txt).strip() == "":
                continue
            heading = label_for(f)
            parts.append(f"**{heading}**\n\n{str(txt).strip()}")
        description = "\n\n".join(parts)

        rows = []
        for f in TABLE_FIELDS:
            val = v(f)
            if val is None or str(val).strip() == "":
                continue
            rows.append([label_for(f), str(val)])

        meta = {
            "id":            pid,
            "name":          name,
            "description":   description,
            "tableColumns":  ["Field", "Value"],
            "tableRows":     rows,
            "sourcePdf":     f"/data/{pid}/source.pdf",
            "aiAnalysisPdf": f"/data/{pid}/ai-analysis.pdf",
        }

        out = ROOT / "data" / pid / "meta.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(meta, indent=2, ensure_ascii=False) + "\n")
        print(f"  wrote {out}  desc={len(description)} chars  rows={len(rows)}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
